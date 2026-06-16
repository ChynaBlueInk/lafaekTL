#!/usr/bin/env python3
"""
Export English strings from public-facing pages that have NO bilingual system yet.
Creates a workbook for translators to fill in the Tetun column.

One sheet per page. Each row: Context | English | Tetun (blank)

Run from the project root:
    python scripts/export-untranslated.py

Output: scripts/translations-needed.xlsx
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

ROOT    = Path(__file__).resolve().parent.parent
APP_DIR = ROOT / "app"
OUT     = ROOT / "scripts" / "translations-needed.xlsx"

# Pages to include — public-facing, genuinely English-only
# Paths relative to APP_DIR
# Excluded: feedback (already has combined "English / Tetun" labels)
#           stories/impact/submit (already mixes both languages inline)
PUBLIC_PAGES = [
    "services/page.tsx",
    "cyber/children/game/page.tsx",
    "magazines/apply/page.tsx",
    "reports/page.tsx",
]

# ── string filtering ──────────────────────────────────────────────────────────

# Patterns that indicate a string is NOT user-facing text
_SKIP_RE = re.compile(
    r'^/'                               # URL / route path
    r'|^https?://'                      # absolute URL
    r'|^#[0-9a-fA-F]{3,8}$'            # hex colour
    r'|^\s*$'                           # blank
    r'|^[a-z][a-zA-Z0-9]*$'            # single camelCase word (identifier)
    r'|^[a-z][a-z0-9-]+$'              # single kebab-case word
    r'|^[A-Z_][A-Z0-9_]{2,}$'         # CONSTANT_NAME or KEY
    r'|^\d[\d.,\s%]+$'                 # pure number / percentage
    r'|\.tsx?$|\.css$|\.svg$|\.png$'   # file extensions
    r'|^@/'                             # internal import alias
    r'|_v\d+$'                          # storage key versioned suffix e.g. cyberHero_v2
    r'|^[a-z][a-zA-Z0-9]+_v\d'        # camelCase + version
    r'|^lafaek_'                        # internal storage key prefix
)

# Variable names whose string values are never user-facing text
_SKIP_KEYS = {
    "className", "style", "src", "href", "alt", "id", "type", "name",
    "key", "ref", "aria-label", "placeholder", "autoComplete", "target",
    "rel", "accept", "method", "action", "encType", "pattern",
    "minLength", "maxLength", "rows", "cols", "STORAGE_KEY", "HERO_KEY",
    "PROGRESS_KEY", "KEY", "variant", "size", "color", "icon",
    "severity", "theme", "slug", "tag", "category", "status",
    "role", "format", "mode", "layout", "shape", "align", "direction",
    # CSS property names
    "fontSize", "fontWeight", "fontFamily", "lineHeight", "letterSpacing",
    "padding", "paddingTop", "paddingBottom", "paddingLeft", "paddingRight",
    "margin", "marginTop", "marginBottom", "marginLeft", "marginRight",
    "border", "borderRadius", "borderTop", "borderBottom", "borderLeft", "borderRight",
    "borderColor", "borderWidth", "borderStyle",
    "background", "backgroundColor", "backgroundImage",
    "boxShadow", "textShadow", "outline",
    "width", "height", "minWidth", "maxWidth", "minHeight", "maxHeight",
    "gap", "rowGap", "columnGap", "gridTemplateColumns", "gridTemplateRows",
    "display", "flexDirection", "justifyContent", "alignItems", "flexWrap",
    "position", "top", "bottom", "left", "right", "zIndex", "overflow",
    "transition", "transform", "animation", "opacity",
    "sm", "md", "lg", "xl",
}

# CSS value patterns — skip these regardless of key name
_CSS_VALUE_RE = re.compile(
    r'^\d+(\.\d+)?(rem|em|px|vh|vw|%|fr)(\s|$)'   # dimension
    r'|^rgba?\s*\('                                  # color function
    r'|^hsl\s*\('
    r'|^linear-gradient'                             # gradient
    r'|^radial-gradient'
    r'|^\d+px\s+\d+px'                              # shadow shorthand
    r'|^solid\b|^dashed\b|^dotted\b'               # border style
    r'|^#[0-9a-fA-F]{3,8}$'                        # hex colour
    r'|^(flex|grid|block|inline|none|auto|normal|bold|center|left|right|top|bottom|middle|wrap|nowrap|hidden|visible|absolute|relative|fixed|sticky)$'
)

def _looks_like_text(s: str) -> bool:
    s = s.strip()
    if len(s) < 4:
        return False
    if _SKIP_RE.search(s):
        return False
    if _CSS_VALUE_RE.search(s):
        return False
    # Must have at least one letter
    if not re.search(r'[a-zA-Z]', s):
        return False
    # Skip Tailwind className blobs (many space-separated hyphenated tokens)
    if re.match(r'^[a-z][a-z0-9-]+([ ][a-z][a-z0-9-/[\]:.!@#$%^&*()]+){3,}$', s):
        return False
    # Skip strings that look like combined English/Tetun (already bilingual)
    if " / " in s and re.search(r'[a-z]{4,}', s.split(" / ")[-1]):
        return False
    return True


# ── extraction strategies ─────────────────────────────────────────────────────

def skip_string(text: str, i: int) -> int:
    q = text[i]
    i += 1
    while i < len(text):
        if text[i] == "\\" and i + 1 < len(text):
            i += 2
            continue
        if text[i] == q:
            return i + 1
        i += 1
    return i


def _unescape(s: str) -> str:
    return (s.replace('\\"', '"').replace("\\'", "'")
             .replace("\\n", " ").replace("\\t", " ")
             .replace("\\\\", "\\"))


# 1. JSX text nodes — literal text between > and <
_JSX_TEXT = re.compile(r'>\s*([^<>{}\n]{4,}?)\s*<')

def extract_jsx_text(text: str) -> list:
    results = []
    for m in _JSX_TEXT.finditer(text):
        s = m.group(1).strip()
        if _looks_like_text(s) and "{" not in s:
            results.append(("JSX text", s))
    return results


# 2. String literals assigned to named variables / object keys
_STRING_ASSIGN = re.compile(
    r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s*(?:=|:)\s*'
    r'("(?:[^"\\]|\\.){4,}"|\'(?:[^\'\\]|\\.){4,}\')'
)

def extract_string_assignments(text: str) -> list:
    results = []
    for m in _STRING_ASSIGN.finditer(text):
        key = m.group(1)
        # skip keys that look like html attrs / imports / css
        if key in _SKIP_KEYS:
            continue
        raw = m.group(2)[1:-1]
        s = _unescape(raw)
        if _looks_like_text(s):
            results.append((key, s))
    return results


# 3. Multi-word string literals used in arrays / JSX expressions
# Requires at least one space to avoid picking up identifiers and code fragments.
_ARRAY_STR = re.compile(
    r'"((?:[^"\\]|\\.){4,})"'
    r'|\'((?:[^\'\\]|\\.){4,})\''
)

def extract_array_strings(text: str) -> list:
    """Catch multi-word string literals not already caught by assignments."""
    results = []
    for m in _ARRAY_STR.finditer(text):
        raw = m.group(1) or m.group(2)
        s = _unescape(raw)
        # Require a space — keeps phrases, drops identifiers and code fragments
        if " " not in s:
            continue
        if _looks_like_text(s):
            results.append(("string literal", s))
    return results


def process_file(path: Path) -> list:
    text = path.read_text(encoding="utf-8", errors="replace")

    rows = []
    seen = set()

    def add(context: str, value: str):
        v = value.strip()
        if v and v not in seen and _looks_like_text(v):
            seen.add(v)
            rows.append((context, v))

    # Strategy 1: named key assignments (most reliable — includes label, desc, title, etc.)
    for ctx, s in extract_string_assignments(text):
        add(ctx, s)

    # Strategy 2: JSX text nodes (hardcoded text between tags)
    for ctx, s in extract_jsx_text(text):
        add(ctx, s)

    # Strategy 3: multi-word string literals for simple pages only
    # Skipped for pages that use heavy JS/game logic (too noisy)
    game_like = len(re.findall(r'\bconst\b', text)) > 30
    if not game_like:
        for ctx, s in extract_array_strings(text):
            add(ctx, s)

    return rows


# ── spreadsheet ───────────────────────────────────────────────────────────────

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="6D3A7A")   # purple — different from review sheet
WRAP     = Alignment(wrap_text=True, vertical="top")
TOP      = Alignment(vertical="top")


def sheet_name(rel_path: str) -> str:
    parts = [p for p in Path(rel_path).parts if p not in ("page.tsx",)]
    name  = "/".join(parts)
    name  = re.sub(r'[\\\/*?:\[\]]', "-", name)
    return name[-31:] if len(name) > 31 else name


def main():
    wb = Workbook()
    wb.remove(wb.active)

    for rel in PUBLIC_PAGES:
        path = APP_DIR / rel
        if not path.exists():
            print(f"  MISSING: {rel}")
            continue

        rows = process_file(path)
        if not rows:
            print(f"  No strings found: {rel}")
            continue

        name = sheet_name(rel)
        ws   = wb.create_sheet(title=name)

        ws.append(["Context / Key", "English", "Tetun (to fill in)", "Notes"])
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = WRAP
        ws.freeze_panes = "A2"

        for ctx, en in rows:
            ws.append([ctx, en, "", ""])
            r = ws.max_row
            ws.cell(r, 1).alignment = TOP
            ws.cell(r, 2).alignment = WRAP
            ws.cell(r, 3).alignment = WRAP
            ws.cell(r, 4).alignment = TOP

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 20

        print(f"  {name}: {len(rows)} strings")

    if not wb.sheetnames:
        print("No strings extracted.")
        return

    wb.save(OUT)
    print(f"\nWritten: {OUT}")


if __name__ == "__main__":
    main()
