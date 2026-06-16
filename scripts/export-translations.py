#!/usr/bin/env python3
"""
Export English / Tetun translation pairs from all app/page.tsx files
to an Excel workbook for translator review.

One sheet per page. Missing Tetun cells are highlighted pink.
Uncertain cells (containing ?) are highlighted yellow.

Run from the project root:
    python scripts/export-translations.py

Output: scripts/translations-review.xlsx
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

ROOT     = Path(__file__).resolve().parent.parent
APP_DIR  = ROOT / "app"
OUT_FILE = ROOT / "scripts" / "translations-review.xlsx"

# ── low-level text helpers ────────────────────────────────────────────────────

def skip_string(text: str, i: int) -> int:
    """i points at an opening quote; return index after the closing quote."""
    q = text[i]
    i += 1
    while i < len(text):
        if text[i] == "\\" and i + 1 < len(text):
            i += 2
            continue
        if text[i] == q:
            return i + 1
        i += 1
    return i


def extract_block(text: str, start: int):
    """
    start points at '{'.  Walk forward matching braces (skipping strings).
    Returns (inner_content, index_after_closing_brace) or (None, start).
    """
    if start >= len(text) or text[start] != "{":
        return None, start
    depth = 0
    i = start
    while i < len(text):
        c = text[i]
        if c in ('"', "'", "`"):
            i = skip_string(text, i)
            continue
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return text[start + 1 : i], i + 1
        i += 1
    return None, start


def extract_array(text: str, start: int):
    """
    start points at '['.  Return (list_of_string_items, index_after_']').
    Only captures top-level string literals (depth == 1 inside the array).
    """
    if start >= len(text) or text[start] != "[":
        return [], start
    depth = 0
    i = start
    items = []
    while i < len(text):
        c = text[i]
        if c == "[":
            depth += 1
            i += 1
            continue
        if c == "]":
            depth -= 1
            if depth == 0:
                return items, i + 1
            i += 1
            continue
        if c in ('"', "'", "`") and depth == 1:
            end = skip_string(text, i)
            raw = text[i + 1 : end - 1]
            items.append(_unescape(raw))
            i = end
            continue
        i += 1
    return items, i


def _unescape(s: str) -> str:
    return (
        s.replace('\\"', '"')
         .replace("\\'", "'")
         .replace("\\n", "\n")
         .replace("\\t", "\t")
         .replace("\\\\", "\\")
    )

# ── find the main translation object ─────────────────────────────────────────

_TRANS_PATTERNS = [
    # const TRANSLATIONS: Record<Lang, T> = {
    r"const\s+TRANSLATIONS\b[^=]*=\s*\{",
    # const t = { en: {...}, tet: {...} }[lang]
    r"const\s+t\s*=\s*\{",
]

def find_main_object(text: str):
    """Return the inner content of the main bilingual translation object, or None."""
    for pat in _TRANS_PATTERNS:
        m = re.search(pat, text)
        if not m:
            continue
        brace_pos = text.rfind("{", m.start(), m.end() + 1)
        if brace_pos == -1:
            continue
        content, _ = extract_block(text, brace_pos)
        if content and re.search(r'\ben\s*:', content) and re.search(r'\btet\s*:', content):
            return content
    return None


def find_lang_block(container: str, lang: str):
    """Find the en: or tet: block inside a container string. Returns inner content or None."""
    pattern = re.compile(r'\b' + lang + r'\s*:\s*\{')
    for m in pattern.finditer(container):
        brace_pos = container.find("{", m.start())
        content, _ = extract_block(container, brace_pos)
        if content is not None:
            return content
    return None

# ── extract key→value pairs from a lang block ─────────────────────────────────

def _parse_pairs(block: str, prefix: str = "") -> list:
    """
    Walk a translation block and yield (full_key, value) pairs where value is
    either a str or a list[str].
    """
    pairs = []
    i = 0
    while i < len(block):
        # skip whitespace / commas / semicolons
        while i < len(block) and block[i] in (" ", "\t", "\n", "\r", ",", ";"):
            i += 1
        if i >= len(block):
            break

        # skip string literals that appear as values or template expressions
        if block[i] in ('"', "'", "`"):
            i = skip_string(block, i)
            continue

        # try to match an identifier key
        km = re.match(r"([a-zA-Z_$][a-zA-Z0-9_$]*)\s*:", block[i:])
        if not km:
            i += 1
            continue

        key = km.group(1)
        full_key = f"{prefix}.{key}" if prefix else key
        i += km.end()

        # skip whitespace after colon
        while i < len(block) and block[i] in (" ", "\t", "\n", "\r"):
            i += 1
        if i >= len(block):
            break

        c = block[i]

        if c in ('"', "'", "`"):
            end = skip_string(block, i)
            value = _unescape(block[i + 1 : end - 1])
            pairs.append((full_key, value))
            i = end

        elif c == "[":
            items, end = extract_array(block, i)
            if items:
                pairs.append((full_key, items))
            else:
                # array of objects — recurse into items to find strings
                # (handled below via object recursion; just skip for now)
                pass
            i = end

        elif c == "{":
            # nested object — recurse
            inner, end = extract_block(block, i)
            if inner:
                sub = _parse_pairs(inner, prefix=full_key)
                pairs.extend(sub)
            i = end

        else:
            # number, boolean, identifier — skip to next comma/newline
            while i < len(block) and block[i] not in (",", "\n", "}"):
                i += 1

    return pairs

# ── bilingual Record<Lang, string> objects in feature/example arrays ──────────

def extract_bilingual_records(text: str) -> list:
    """
    Find objects like { en: "...", tet: "..." } or { en: [...], tet: [...] }
    that appear outside the main TRANSLATIONS block (e.g. Feature / Example arrays).
    Returns list of (key_hint, en_val, tet_val).
    """
    results = []
    # Find property assignments whose value is a bilingual object:
    #   propName: { en: "...", tet: "..." }
    prop_re = re.compile(r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s*:\s*\{')
    for m in prop_re.finditer(text):
        prop = m.group(1)
        brace = text.find("{", m.start())
        inner, _ = extract_block(text, brace)
        if inner is None:
            continue
        # must have both en: and tet: as direct children
        en_m  = re.search(r'\ben\s*:\s*', inner)
        tet_m = re.search(r'\btet\s*:\s*', inner)
        if not en_m or not tet_m:
            continue

        def read_value(block, after):
            j = after
            while j < len(block) and block[j] in (" ", "\t", "\n", "\r"):
                j += 1
            if j >= len(block):
                return None
            if block[j] in ('"', "'", "`"):
                end = skip_string(block, j)
                return _unescape(block[j + 1 : end - 1])
            if block[j] == "[":
                items, _ = extract_array(block, j)
                return items
            return None

        en_val  = read_value(inner, en_m.end())
        tet_val = read_value(inner, tet_m.end())

        if en_val is None and tet_val is None:
            continue

        if isinstance(en_val, list) or isinstance(tet_val, list):
            en_list  = en_val  if isinstance(en_val,  list) else []
            tet_list = tet_val if isinstance(tet_val, list) else []
            length = max(len(en_list), len(tet_list))
            for idx in range(length):
                ev = en_list[idx]  if idx < len(en_list)  else ""
                tv = tet_list[idx] if idx < len(tet_list) else ""
                results.append((f"{prop}[{idx}]", ev, tv))
        else:
            results.append((prop, en_val or "", tet_val or ""))

    return results


# ── inline ternary strings ────────────────────────────────────────────────────

_TERNARY_TET_FIRST = re.compile(
    r'(?:lang|language)\s*===?\s*["\']tet["\']\s*\?'
    r'\s*("(?:[^"\\]|\\.)*"|\'(?:[^\'\\]|\\.)*\')'
    r'\s*:\s*'
    r'("(?:[^"\\]|\\.)*"|\'(?:[^\'\\]|\\.)*\')'
)
_TERNARY_EN_FIRST = re.compile(
    r'(?:lang|language)\s*===?\s*["\']en["\']\s*\?'
    r'\s*("(?:[^"\\]|\\.)*"|\'(?:[^\'\\]|\\.)*\')'
    r'\s*:\s*'
    r'("(?:[^"\\]|\\.)*"|\'(?:[^\'\\]|\\.)*\')'
)

def extract_ternaries(text: str) -> list:
    results = []
    for m in _TERNARY_TET_FIRST.finditer(text):
        tet = _unescape(m.group(1)[1:-1])
        en  = _unescape(m.group(2)[1:-1])
        results.append(("(inline)", en, tet))
    for m in _TERNARY_EN_FIRST.finditer(text):
        en  = _unescape(m.group(1)[1:-1])
        tet = _unescape(m.group(2)[1:-1])
        results.append(("(inline)", en, tet))
    return results

# ── per-file extraction ───────────────────────────────────────────────────────

def process_file(path: Path) -> list:
    """Return list of (key, en, tet) for the page at path."""
    text = path.read_text(encoding="utf-8", errors="replace")

    # Quick filter — skip pages with no bilingual content at all
    if not re.search(r'\btet\b', text):
        return []

    rows: list = []
    seen_en: set = set()  # deduplicate ternaries against TRANSLATIONS content

    # ── Strategy 1: main TRANSLATIONS / t object ──────────────────────────
    main_obj = find_main_object(text)
    if main_obj:
        en_block  = find_lang_block(main_obj, "en")
        tet_block = find_lang_block(main_obj, "tet")

        if en_block and tet_block:
            en_map  = dict(_parse_pairs(en_block))
            tet_map = dict(_parse_pairs(tet_block))

            all_keys = list(en_map.keys())
            for k in tet_map:
                if k not in all_keys:
                    all_keys.append(k)

            for key in all_keys:
                ev = en_map.get(key, "")
                tv = tet_map.get(key, "")

                if isinstance(ev, list) or isinstance(tv, list):
                    en_list  = ev if isinstance(ev,  list) else []
                    tet_list = tv if isinstance(tv, list) else []
                    length = max(len(en_list), len(tet_list))
                    for idx in range(length):
                        e = en_list[idx]  if idx < len(en_list)  else ""
                        t = tet_list[idx] if idx < len(tet_list) else ""
                        rows.append((f"{key}[{idx}]", e, t))
                        seen_en.add(e)
                else:
                    rows.append((key, str(ev) if ev != "" else "", str(tv) if tv != "" else ""))
                    seen_en.add(str(ev))

    # ── Strategy 2: bilingual Record<Lang,string> in feature/example arrays ──
    bil = extract_bilingual_records(text)
    existing_keys = {r[0] for r in rows}
    for key, ev, tv in bil:
        if key not in existing_keys:
            rows.append((key, ev, tv))
            seen_en.add(ev)

    # ── Strategy 3: inline ternaries ──────────────────────────────────────
    for key, ev, tv in extract_ternaries(text):
        if ev not in seen_en:
            rows.append((key, ev, tv))
            seen_en.add(ev)

    return rows

# ── spreadsheet output ────────────────────────────────────────────────────────

HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL    = PatternFill("solid", fgColor="1A5276")
MISS_FILL   = PatternFill("solid", fgColor="FFD6D6")   # pink  — missing Tetun
UNSURE_FILL = PatternFill("solid", fgColor="FFF3CD")   # amber — contains ?
WRAP        = Alignment(wrap_text=True, vertical="top")
TOP         = Alignment(vertical="top")


def sheet_name(path: Path) -> str:
    rel   = path.relative_to(APP_DIR)
    parts = [p for p in rel.parts if p not in ("page.tsx", "layout.tsx")]
    name  = "/".join(parts) if parts else "root"
    # Excel limit: 31 chars, no special chars
    name  = re.sub(r'[\\\/*?:\[\]]', "-", name)
    return name[-31:] if len(name) > 31 else name


def write_workbook(pages: dict):
    wb = Workbook()
    wb.remove(wb.active)

    for name, rows in pages.items():
        ws = wb.create_sheet(title=name)

        # header row
        ws.append(["Key", "English", "Tetun", "Translator notes"])
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = WRAP
        ws.freeze_panes = "A2"

        # data rows
        for key, en, tet in rows:
            ws.append([key, en, tet, ""])
            r = ws.max_row
            ws.cell(r, 1).alignment = TOP
            ws.cell(r, 2).alignment = WRAP
            ws.cell(r, 3).alignment = WRAP
            ws.cell(r, 4).alignment = TOP

            if not tet.strip():
                ws.cell(r, 3).fill = MISS_FILL
            elif "?" in tet:
                ws.cell(r, 3).fill = UNSURE_FILL

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 22

    wb.save(OUT_FILE)


def main():
    pages = {}
    skipped = []

    for tsx in sorted(APP_DIR.rglob("page.tsx")):
        rows = process_file(tsx)
        if rows:
            pages[sheet_name(tsx)] = rows
        else:
            skipped.append(tsx.relative_to(ROOT))

    if not pages:
        print("No bilingual content found under app/.")
        return

    write_workbook(pages)

    print(f"Written: {OUT_FILE}")
    print(f"  Sheets ({len(pages)}): {', '.join(pages)}")
    if skipped:
        print(f"  Skipped (no bilingual content): {', '.join(str(p) for p in skipped)}")


if __name__ == "__main__":
    main()
